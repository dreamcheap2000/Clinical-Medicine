#!/usr/bin/env python3
"""
build_edu_data.py
-----------------
Scans the 'Patient education' folder, converts .docx files to HTML
using mammoth, and writes PHCEP/data/edu/patient_edu_data.json
using the v2.0 schema (with FastSR SOAP sections and multi-language versions).

When OPENAI_API_KEY is set (e.g., via GitHub Actions secret), the script calls
the OpenAI API to:
  - Translate professional content into the other 2 language versions
  - Extract source URL(s) from the text
  - Auto-generate FastSR S/O/A/P classification

Existing manually-crafted entries (those NOT matching any file in the folder)
are preserved unchanged.

Supported file types:
  .docx  → converted to HTML with mammoth
  .txt   → plain text, wrapped in <p> tags
  .xlsx  → workbook tables rendered to HTML
  other  → stored with a direct GitHub raw URL
"""

import json
import os
import re
import datetime
from html import escape
from pathlib import Path

REPO_ROOT   = Path(__file__).resolve().parents[2]
EDU_FOLDER  = REPO_ROOT / "Patient education"
OUTPUT_FILE = REPO_ROOT / "PHCEP" / "data" / "edu" / "patient_edu_data.json"

GITHUB_RAW_BASE = (
    "https://raw.githubusercontent.com/dreamcheap2000/Clinical-Medicine/main/"
    "Patient%20education/"
)

SECTION_MARKER_RE = re.compile(r"^\s*([@#&])\s*(\d+)?(?:\s+(.+?))?\s*$")


def docx_to_html(docx_path: Path) -> tuple[str, str]:
    """Convert a .docx file to HTML using mammoth.
    Returns (html_string, plain_text).
    """
    try:
        import mammoth
    except ImportError as exc:
        raise ImportError(
            "mammoth is required. Install with: pip install mammoth"
        ) from exc

    style_map = """
p[style-name='Heading 1'] => h2:fresh
p[style-name='Heading 2'] => h3:fresh
p[style-name='Heading 3'] => h4:fresh
"""
    result = mammoth.convert_to_html(docx_path, style_map=style_map)
    # Also extract raw text for FastSR classification and language detection
    text_result = mammoth.extract_raw_text(docx_path)
    return result.value, text_result.value


def txt_to_html(txt_path: Path) -> tuple[str, str]:
    """Read a plain text file and wrap paragraphs in <p> tags."""
    raw = txt_path.read_text(encoding="utf-8", errors="replace")
    html = plain_text_to_html(raw)
    return html, raw


def plain_text_to_html(raw: str) -> str:
    """Convert plain text into paragraph-oriented HTML."""
    paragraphs = [p.strip() for p in raw.split("\n\n") if p.strip()]
    return "".join(
        f"<p>{escape(p).replace(chr(10), '<br>')}</p>"
        for p in paragraphs
    )


def xlsx_to_html(xlsx_path: Path) -> tuple[str, str]:
    """Render workbook sheets as HTML tables and plain text."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required. Install with: pip install openpyxl"
        ) from exc

    wb = load_workbook(xlsx_path, data_only=True)
    html_parts: list[str] = []
    plain_parts: list[str] = []

    for ws in wb.worksheets:
        rows: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            vals = ["" if v is None else str(v).strip() for v in row]
            last_non_empty = max(
                (i for i, v in enumerate(vals) if v),
                default=-1,
            )
            if last_non_empty < 0:
                continue
            rows.append(vals[: last_non_empty + 1])

        if not rows:
            continue

        html_parts.append(f"<h3>{escape(ws.title)}</h3>")
        plain_parts.append(ws.title)

        if len(rows) == 1:
            html_parts.append(f"<p>{escape(' | '.join(rows[0]))}</p>")
            plain_parts.append(" | ".join(rows[0]))
            continue

        header = rows[0]
        body = rows[1:]
        html_parts.append("<table><thead><tr>")
        html_parts.extend(f"<th>{escape(cell)}</th>" for cell in header)
        html_parts.append("</tr></thead><tbody>")
        for row in body:
            padded = row + [""] * max(0, len(header) - len(row))
            html_parts.append("<tr>")
            html_parts.extend(f"<td>{escape(cell)}</td>" for cell in padded[:len(header)])
            html_parts.append("</tr>")
            plain_parts.append(" | ".join(cell for cell in padded[:len(header)] if cell))
        html_parts.append("</tbody></table>")

    return "".join(html_parts), "\n".join(plain_parts)


def strip_leading_title_para(html: str, stem: str) -> str:
    """Remove any leading <p>/<h1>/<h2> that is an exact duplicate of the document title/stem.

    Docx files often include the document title as the first heading/paragraph, which
    duplicates the ``title`` field shown in the viewer header.  This function strips that
    redundant first element so it does not appear in the rendered content.
    """
    stripped = stem.strip()
    if not stripped or not html:
        return html
    # Match a leading block-level tag whose text content equals the stem exactly.
    pattern = re.compile(
        r'^\s*<(p|h1|h2|h3)(?:\s[^>]*)?>\s*'
        + re.escape(stripped)
        + r'\s*</\1>\s*',
        re.IGNORECASE,
    )
    return pattern.sub("", html).strip()


def sanitize_id(filename: str, idx: int) -> str:
    stem = Path(filename).stem
    stem_ascii = re.sub(r"[^a-zA-Z0-9_\-]", "_", stem)[:30]
    return f"edu{idx+1:03d}" if not stem_ascii else f"edu{idx+1:03d}_{stem_ascii}"


def sanitize_section_id(filename: str, marker: str, idx: int) -> str:
    stem = Path(filename).stem
    marker_ascii = re.sub(r"[^a-zA-Z0-9_\-]", "_", marker).strip("_")
    base = re.sub(r"[^a-zA-Z0-9_\-]", "_", f"{stem}_{marker_ascii}")[:40]
    return f"edu{idx+1:03d}" if not base else f"edu{idx+1:03d}_{base}"


def load_existing() -> dict:
    """Load existing patient_edu_data.json; return empty v2 dict on failure."""
    if OUTPUT_FILE.exists():
        try:
            return json.loads(OUTPUT_FILE.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {"version": "2.0", "generated": "", "entries": []}


SIDECAR_SUFFIX = ".meta.json"


def load_sidecar(fpath: Path) -> dict:
    """Load an optional sidecar <filename>.meta.json file next to the document.

    The sidecar may contain:
      source_urls  – list of URLs (merged with any URLs found in the document)
      title        – override title string
      tags         – list of tag strings
      notes        – freeform notes (ignored by the pipeline)

    Returns an empty dict if no sidecar exists or parsing fails.
    """
    sidecar = fpath.with_name(fpath.name + SIDECAR_SUFFIX)
    if not sidecar.exists():
        return {}
    try:
        return json.loads(sidecar.read_text(encoding="utf-8"))
    except Exception as exc:
        print(f"  ⚠️ Could not read sidecar {sidecar.name}: {exc}")
        return {}


def normalize_marker(raw_marker: str, fallback_index: int | None = None) -> tuple[str, str]:
    """Normalize @/#/& markers to a stable @N form and return title hint."""
    match = SECTION_MARKER_RE.match(raw_marker.strip())
    if not match:
        marker = raw_marker.strip()
        return marker, ""
    symbol, number, title_hint = match.groups()
    symbol = "@" if symbol in {"#", "&"} else symbol
    number = number or (str(fallback_index) if fallback_index is not None else "")
    marker = f"{symbol}{number}" if number else symbol
    return marker, (title_hint or "").strip()


def get_sidecar_sections(sidecar: dict) -> dict[str, dict]:
    """Return section metadata keyed by normalized marker."""
    sections: dict[str, dict] = {}
    for idx, item in enumerate(sidecar.get("sections", []), start=1):
        marker, _ = normalize_marker(str(item.get("marker", f"@{idx}")), idx)
        sections[marker] = item
    return sections


def split_marked_text(text: str) -> list[dict]:
    """Split marker-delimited text into sections."""
    sections: list[dict] = []
    current: dict | None = None
    auto_index = 0

    for raw_line in text.splitlines():
        stripped = raw_line.strip()
        if SECTION_MARKER_RE.match(stripped):
            raw_match = SECTION_MARKER_RE.match(stripped)
            if raw_match is None:
                raise ValueError(f"Invalid section marker: {stripped}")
            has_explicit_number = bool(raw_match.group(2))
            if not has_explicit_number:
                auto_index += 1
            marker, title_hint = normalize_marker(
                stripped,
                auto_index if not has_explicit_number else None,
            )
            if current and marker == current.get("marker"):
                current["title_hint"] = title_hint or current.get("title_hint", "")
                continue
            if current and any(line.strip() for line in current["lines"]):
                sections.append(current)
            if current and not any(line.strip() for line in current["lines"]):
                current["marker"] = marker
                current["title_hint"] = title_hint or current.get("title_hint", "")
            else:
                current = {"marker": marker, "title_hint": title_hint, "lines": []}
            continue
        if current is not None:
            current["lines"].append(raw_line.rstrip())

    if current and any(line.strip() for line in current["lines"]):
        sections.append(current)

    return [
        {
            "marker": sec["marker"],
            "title_hint": sec.get("title_hint", ""),
            "text": "\n".join(sec["lines"]).strip(),
        }
        for sec in sections
    ]


def is_link_only_entry(entry: dict | None) -> bool:
    """Return True when an entry is only a bare download-link placeholder."""
    if not entry:
        return True
    versions = entry.get("versions") or {}
    simple = (versions.get("simple_zh") or "").strip()
    professional = (versions.get("professional_zh") or "").strip()
    english = (versions.get("english") or "").strip()
    if english:
        return False
    combined = " ".join(v for v in [simple, professional] if v)
    if not combined:
        return True
    text_only = re.sub(r"<[^>]+>", " ", combined)
    text_only = re.sub(r"\s+", " ", text_only).strip()
    link_count = combined.count("href=")
    return link_count >= 1 and len(text_only.replace("📎", "").strip()) <= 18


def _needs_ai(versions: dict) -> bool:
    """Return True if any language version is missing or incomplete."""
    return not (
        versions.get("simple_zh", "").strip()
        and versions.get("professional_zh", "").strip()
        and versions.get("english", "").strip()
    )


def build():
    if not EDU_FOLDER.exists():
        print(f"ERROR: Folder not found: {EDU_FOLDER}")
        return

    # Import translation helper
    import sys
    sys.path.insert(0, str(Path(__file__).parent))
    from translate_edu import process_document, build_prototypes, build_fastsr

    existing_data = load_existing()
    existing_entries = existing_data.get("entries") or []

    # Build a lookup: title → existing entry (for preservation)
    existing_by_title = {e.get("title", ""): e for e in existing_entries}
    # Build a lookup: filename stem → existing entry (for update matching)
    existing_by_stem: dict[str, dict] = {}
    for e in existing_entries:
        src = e.get("source_file", "")
        if src:
            existing_by_stem[Path(src).stem] = e
    existing_by_filename_title = {
        (e.get("source_file", ""), e.get("title", "")): e
        for e in existing_entries
        if e.get("source_file") and e.get("title")
    }
    existing_by_source_label = {
        e.get("source_label", ""): e
        for e in existing_entries
        if e.get("source_label")
    }

    file_list = sorted(
        f for f in EDU_FOLDER.iterdir()
        if f.is_file() and not f.name.startswith(".")
    )

    new_entries: list[dict] = []
    processed_titles: set[str] = set()

    for idx, fpath in enumerate(file_list):
        suffix = fpath.suffix.lower()
        filename = fpath.name
        stem = fpath.stem

        # Skip sidecar metadata files — they are helpers, not documents
        if filename.endswith(SIDECAR_SUFFIX):
            continue

        print(f"Processing [{suffix}]: {filename}")

        # Load optional sidecar metadata
        sidecar = load_sidecar(fpath)
        extra_urls: list[str] = sidecar.get("source_urls", [])
        title_override: str = sidecar.get("title", "")
        extra_tags: list[str] = sidecar.get("tags", [])
        sidecar_sections = get_sidecar_sections(sidecar)

        existing_entry = existing_by_stem.get(stem) or existing_by_title.get(stem)

        if suffix == ".docx":
            html_content, plain_text = docx_to_html(fpath)
            # Strip redundant leading title paragraph (docx often includes it as first element)
            html_content = strip_leading_title_para(html_content, stem)
        elif suffix == ".txt":
            html_content, plain_text = txt_to_html(fpath)
            html_content = strip_leading_title_para(html_content, stem)
        elif suffix == ".xlsx":
            html_content, plain_text = xlsx_to_html(fpath)
            merged_tags = list(dict.fromkeys(((existing_entry or {}).get("tags") or []) + extra_tags))
            title = title_override or (existing_entry or {}).get("title") or stem
            fastsr = (
                (existing_entry or {}).get("fastsr")
                or build_fastsr(plain_text[:12000])
            )
            source_urls = list(dict.fromkeys(extra_urls + ((existing_entry or {}).get("source_urls") or [])))
            if not source_urls:
                source_urls = [GITHUB_RAW_BASE + filename.replace(" ", "%20")]
            source_url = source_urls[0]
            existing_versions = (existing_entry or {}).get("versions") or {}
            entry = {
                "id": (existing_entry or {}).get("id") or sanitize_id(filename, idx),
                "title": title,
                "source_file": filename,
                "source_url": source_url,
                "source_label": (existing_entry or {}).get("source_label") or filename,
                "source_urls": source_urls,
                "original_lang": (existing_entry or {}).get("original_lang", "zh-TW"),
                "added_date": (existing_entry or {}).get("added_date") or datetime.date.today().isoformat(),
                "version": (existing_entry or {}).get("version", "1"),
                "tags": merged_tags,
                "fastsr": fastsr,
                "prototype": build_prototypes(
                    title=title,
                    tags=merged_tags,
                    fastsr=fastsr,
                ),
                "versions": {
                    "simple_zh": html_content,
                    "professional_zh": html_content,
                    "english": existing_versions.get("english", ""),
                },
            }
            new_entries.append(entry)
            processed_titles.add(entry["title"])
            continue
        else:
            if suffix == ".pdf" and is_link_only_entry(existing_entry):
                print("  ↷ Skipping PDF link-only placeholder")
                continue
            raw_url = GITHUB_RAW_BASE + filename.replace(" ", "%20")
            merged_tags = list(dict.fromkeys(((existing_entry or {}).get("tags") or []) + extra_tags))
            title = title_override or (existing_entry or {}).get("title") or stem
            fastsr = (existing_entry or {}).get("fastsr") or {"S": [], "O": [], "A": [], "P": []}
            source_urls = list(dict.fromkeys(extra_urls + ((existing_entry or {}).get("source_urls") or [raw_url])))
            source_url = source_urls[0] if source_urls else raw_url
            entry = {
                "id": (existing_entry or {}).get("id") or sanitize_id(filename, idx),
                "title": title,
                "source_file": filename,
                "source_url": source_url,
                "source_label": (existing_entry or {}).get("source_label", ""),
                "source_urls": source_urls,
                "original_lang": (existing_entry or {}).get("original_lang", "zh-TW"),
                "added_date": (existing_entry or {}).get("added_date") or datetime.date.today().isoformat(),
                "version": (existing_entry or {}).get("version", "1"),
                "tags": merged_tags,
                "fastsr": fastsr,
                "prototype": (existing_entry or {}).get("prototype") or build_prototypes(
                    title=title,
                    tags=merged_tags,
                    fastsr=fastsr,
                ),
                "versions": (existing_entry or {}).get("versions") or {
                    "simple_zh": (
                        f'<p>📎 <a href="{raw_url}" target="_blank" rel="noopener">'
                        f'下載 {filename}</a></p>'
                    ),
                    "professional_zh": "",
                    "english": "",
                },
            }
            new_entries.append(entry)
            processed_titles.add(entry["title"])
            continue

        sections = split_marked_text(plain_text) if sidecar_sections else []
        if sections:
            used_existing_ids: set[str] = set()
            for section_idx, section in enumerate(sections, start=1):
                marker = section["marker"]
                label = f"{filename} {marker}"
                section_meta = sidecar_sections.get(marker, {})
                section_tags = list(dict.fromkeys(extra_tags + section_meta.get("tags", [])))
                section_entry = existing_by_source_label.get(label)
                if not section_entry:
                    candidate = existing_by_filename_title.get((filename, section_meta.get("title", "")))
                    if candidate and candidate.get("id") not in used_existing_ids:
                        section_entry = candidate
                section_text = section["text"]
                section_html = plain_text_to_html(section_text)
                section_existing_title = (
                    section_meta.get("title")
                    or section.get("title_hint")
                    or (section_entry.get("title") if section_entry else None)
                )
                section_doc = process_document(
                    section_text,
                    section_html,
                    f"{stem}_{marker.lstrip('@')}",
                    existing_title=section_existing_title,
                    extra_urls=extra_urls,
                    existing_versions=(section_entry or {}).get("versions"),
                    existing_fastsr=(section_entry or {}).get("fastsr"),
                )
                entry = {
                    "id": (section_entry or {}).get("id") or sanitize_section_id(filename, marker, idx + section_idx),
                    "title": section_doc["title"],
                    "source_file": filename,
                    "source_url": section_doc["source_url"],
                    "source_label": label,
                    "source_urls": section_doc.get("source_urls", []),
                    "original_lang": (section_entry or {}).get("original_lang") or ("zh-TW" if section_doc["versions"]["professional_zh"] else "en"),
                    "added_date": (section_entry or {}).get("added_date") or datetime.date.today().isoformat(),
                    "version": (section_entry or {}).get("version", "1"),
                    "tags": list(dict.fromkeys(((section_entry or {}).get("tags") or []) + section_tags)),
                    "fastsr": section_doc["fastsr"],
                    "prototype": build_prototypes(
                        title=section_doc["title"],
                        tags=list(dict.fromkeys(((section_entry or {}).get("tags") or []) + section_tags)),
                        fastsr=section_doc["fastsr"],
                    ),
                    "versions": section_doc["versions"],
                }
                new_entries.append(entry)
                processed_titles.add(entry["title"])
                used_existing_ids.add(entry["id"])
            continue

        # Check if this file was already processed.
        # existing_by_stem uses the source_file stem as key.
        # existing_by_title uses entry title; for docx-generated entries the
        # title was historically set to the filename stem, so this also matches.
        existing_title = title_override or (existing_entry.get("title") if existing_entry else None)
        existing_versions = (existing_entry or {}).get("versions") or {}

        needs_ai = _needs_ai(existing_versions) or bool(extra_urls)
        if not needs_ai:
            print(f"  ✔ All versions present and no new URLs — skipping AI")

        # Process via AI translation (passing existing versions as fallback)
        doc_info = process_document(
            plain_text,
            html_content,
            stem,
            existing_title=existing_title,
            extra_urls=extra_urls,
            existing_versions=existing_versions,
            existing_fastsr=(existing_entry or {}).get("fastsr"),
        )

        # Determine entry id
        entry_id = (existing_entry or {}).get("id") or sanitize_id(filename, idx)

        # Merge tags: existing + sidecar (deduplicated)
        merged_tags = list(dict.fromkeys(
            ((existing_entry or {}).get("tags") or []) + extra_tags
        ))

        # Merge: keep existing tags / added_date / version if available
        entry = {
            "id": entry_id,
            "title": doc_info["title"],
            "source_file": filename,
            "source_url": doc_info["source_url"],
            "source_label": (existing_entry or {}).get("source_label") or doc_info["source_label"],
            "source_urls": doc_info.get("source_urls", []),
            "original_lang": (existing_entry or {}).get("original_lang") or ("zh-TW" if doc_info["versions"]["professional_zh"] else "en"),
            "added_date": (existing_entry or {}).get("added_date") or datetime.date.today().isoformat(),
            "version": (existing_entry or {}).get("version", "1"),
            "tags": merged_tags,
            "fastsr": doc_info["fastsr"],
            "prototype": build_prototypes(
                title=doc_info["title"],
                tags=merged_tags,
                fastsr=doc_info["fastsr"],
            ),
            "versions": doc_info["versions"],
        }
        new_entries.append(entry)
        processed_titles.add(doc_info["title"])
        if existing_title:
            processed_titles.add(existing_title)

    # Preserve existing manually-crafted entries not matched by any file
    for e in existing_entries:
        title = e.get("title", "")
        source_file = e.get("source_file", "")
        if title not in processed_titles and not source_file:
            # This is a manually-crafted entry with no source file → keep it
            new_entries.append(e)
            print(f"Preserving manual entry: {title}")

    output = {
        "version": "2.0",
        "generated": datetime.date.today().isoformat(),
        "entries": new_entries,
    }

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"\n✅ Written {len(new_entries)} entries to {OUTPUT_FILE}")


if __name__ == "__main__":
    build()
